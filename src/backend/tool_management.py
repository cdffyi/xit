from hashlib import md5
from os import path

from cryptography.fernet import InvalidToken
from openpyxl import Workbook, load_workbook

from src.backend.errors import InterfaceError
from tokens import secrets
from .encryption import encrypt_dict, decrypt_dict


class Tool:
    def __init__(self, file_path=None):
        self.wb = None
        self.name = None
        if file_path is not None:
            self.name = file_path
            self.wb = load_workbook(file_path)
        else:
            self.wb = Workbook()

    @property
    def file_name(self):
        if ".xlsx" in self.name:
            return self.name
        return f"{self.name}.xlsx"

    @property
    def tool_name(self):
        if ".xlsx" not in self.name:
            return self.name
        return self.name[0:-5]

    @property
    def enc_md5(self):
        return md5(f"{secrets['enc_uuid']}".encode("utf-8"))

    @property
    def manifest(self):
        try:
            if "MANIFEST" in self.wb.get_sheet_names():
                return decrypt_dict(self.wb['MANIFEST']['A1'].value)
            return False
        except InvalidToken:
            InterfaceError(
                f"Filed to load manifest for '{self.name}'")
            return False

    def create(self, name, owner, uid, org, dataset, admin_password):
        # add name to class (missing data)
        self.name = name

        # make sure the workbook is not going to get overwritten
        if path.exists(self.file_name):
            raise InterfaceError(f"Workbook {self.file_name} already exists.")

        # check to see if the manifest exists
        if "MANIFEST" in self.wb.get_sheet_names():
            raise InterfaceError(
                "Can't override tool creation on a current tool.")

        # manifest data
        manifest_data = {
            "name": name,
            "owner": owner,
            "uid": uid,
            "password": admin_password,
            "organization": org,
            "dataset": dataset,
            "version_object": (1, 0, 0),
            "ready": False
        }

        # create manifest sheet
        manifest = self.wb.active
        manifest.title = "MANIFEST"

        # tool manifest encrypted
        manifest['A1'] = encrypt_dict(manifest_data)

        manifest.protection.set_password(admin_password, already_hashed=False)
        manifest.protection.enable()

        # hide sheet
        manifest.sheet_state = 'hidden'

        # create new sheets for tools
        if dataset == "xy":
            dataset_sheet = self.wb.create_sheet("DATASET")
            dataset_sheet['A1'] = "x"
            dataset_sheet['B1'] = "y"

        # change meta
        self.wb.properties.title = name
        self.wb.properties.creator = f"{owner} via XIT"
        self.wb.properties.version_object = (1, 0, 0)

        # check name and save
        self.save()
        return True

    def save(self):
        try:
            self.wb.save(self.file_name)
        except PermissionError:
            raise InterfaceError(
                f"The workbook {self.file_name} is being used by another program, please close the other program and "
                f"try again.")
        except Exception as e:
            print(e)
            raise InterfaceError(
                f"could not save {self.file_name}. fatal error, make sure it is not open in another app.")

# def get_manifest(file_path, password):
#     wb = load_workbook(filename=file_path, data_only=True, read_only=False)
#     save(wb, file_path)
#
#     if "MANIFEST" not in wb.sheetnames:
#         raise InterfaceError("The excel workbook you selected is not a valid tool.")
#
#     manifest = bytes(wb["MANIFEST"]['A1'].value.encode())
#
#     return loads(decrypt(manifest, password))
#
# def load_manifest(file_path, password):
#
